import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

class ReportGenerator:
    def __init__(self, report_path="data/reports/"):
        self.report_path = report_path
        import os
        os.makedirs(report_path, exist_ok=True)

    def generate(self, data: str) -> str:
        # 生成Excel报告
        wb = Workbook()
        ws = wb.active
        ws.title = "审核报告"

        # 标题
        ws["A1"] = "生产材料审核报告"
        ws["A1"].font = Font(size=16, bold=True)
        ws.merge_cells("A1:D1")
        ws["A1"].alignment = Alignment(horizontal="center")

        # 数据写入
        ws.append(["审核项", "内容", "状态", "备注"])
        ws.append(["材料信息", data, "已审核", "AI自动生成"])

        report_file = f"{self.report_path}/audit_report.xlsx"
        wb.save(report_file)
        return report_file